import argparse
import json
import os
from pathlib import Path

import openpyxl


def load_config(config_path: Path):
    with config_path.open("r", encoding="utf-8") as f:
        cfg = json.load(f)
    controllers = cfg.get("network", {}).get("controllers", [])
    ip_to_index = {str(c.get("ip")).strip(): i for i, c in enumerate(controllers) if c.get("ip")}
    start_universe_by_index = {i: int(c.get("startUniverse", 0) or 0) for i, c in enumerate(controllers)}
    channels_per_led = int(cfg.get("mapping", {}).get("channelsPerLed", 3) or 3)
    return ip_to_index, start_universe_by_index, channels_per_led


def main():
    ap = argparse.ArgumentParser(description="Convert prof Ecran.xlsx (eHuB sheet) to led_wall_mapping.csv")
    ap.add_argument("--xlsx", required=True, help="Path to Ecran.xlsx")
    ap.add_argument("--out", required=True, help="Output CSV path (led_wall_mapping.csv)")
    ap.add_argument("--config", default="Assets/StreamingAssets/config.json", help="PixelHub config.json")
    ap.add_argument("--sheet", default="eHuB", help="Sheet name (default: eHuB)")
    ap.add_argument("--header-row", type=int, default=1, help="Header row index (1-based)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    config_path = Path(args.config)

    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")
    if not config_path.exists():
        raise SystemExit(f"config.json not found: {config_path}")

    ip_to_index, start_universe_by_index, channels_per_led = load_config(config_path)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[args.sheet]

    # Expected columns (from your dump):
    # Name | Entity Start | Entity End | ArtNet IP | ArtNet Universe | ...
    # We'll locate them by header names to be robust.
    header = [ws.cell(args.header_row, c).value for c in range(1, 25)]
    header_norm = [str(h).strip().lower() if h is not None else "" for h in header]

    def col(name: str) -> int:
        name = name.lower()
        for i, h in enumerate(header_norm):
            if h == name:
                return i + 1
        raise KeyError(f"Column '{name}' not found in header row {args.header_row}: {header}")

    c_entity_start = col("entity start")
    c_entity_end = col("entity end")
    c_ip = col("artnet ip")
    c_universe = col("artnet universe")

    out_lines = ["entityId,controllerIndex,universe,channel"]
    kept = 0

    # Data starts after header row
    for r in range(args.header_row + 1, ws.max_row + 1):
        entity_start = ws.cell(r, c_entity_start).value
        entity_end = ws.cell(r, c_entity_end).value
        ip = ws.cell(r, c_ip).value
        universe = ws.cell(r, c_universe).value

        if entity_start is None or entity_end is None or ip is None or universe is None:
            continue

        try:
            entity_start = int(entity_start)
            entity_end = int(entity_end)
            universe = int(universe)
            ip = str(ip).strip()
        except Exception:
            continue

        if ip not in ip_to_index:
            # Skip unknown controllers to keep config-driven mapping
            continue

        controller_index = ip_to_index[ip]
        controller_start_universe = start_universe_by_index.get(controller_index, 0)
        absolute_universe = controller_start_universe + universe
        # Each entity maps sequentially starting at DMX channel 0
        for entity_id in range(entity_start, entity_end + 1):
            channel0 = (entity_id - entity_start) * channels_per_led
            if channel0 + (channels_per_led - 1) >= 512:
                # If a range would overflow an universe, stop (shouldn't happen with provided ranges)
                break
            out_lines.append(f"{entity_id},{controller_index},{absolute_universe},{channel0}")
            kept += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")

    print(f"controllers_in_config={len(ip_to_index)} channelsPerLed={channels_per_led}")
    print(f"wrote={out_path} entities={kept}")


if __name__ == "__main__":
    main()

